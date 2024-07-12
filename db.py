import sqlite3
from datetime import datetime, timedelta
import requests
import openpyxl
import re
from openpyxl import Workbook

# Функция для создания таблицы в базе данных
def create_table():
    conn = sqlite3.connect('vacancies_new.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS vacancies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_title TEXT,
            job_url TEXT,
            company_name TEXT,
            source_url TEXT,
            added_date TEXT,
            employer_posted_date TEXT,
            UNIQUE(job_url)
        )
    ''')
    conn.commit()
    conn.close()

def convert_posted_date(posted_date):
    posted_date=posted_date.replace('30+','31')
    return (datetime.now() - timedelta(days=int(posted_date.split()[1]))).strftime('%Y-%m-%d')

# Функция для добавления вакансии в базу данных
def add_vacancy(job_title, job_url, company_name, source_url, employer_posted_date):
    added_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if('-' not in employer_posted_date):
        try:
            employer_posted_date = convert_posted_date(employer_posted_date)
        except:
            pass
    # try:
    #     r=requests.get(job_url).status_code
    #     if(r==404 or r == '404'):
    #         return
    # except:
    #     pass
    company_name = company_name.replace('Careers at ','')

    split_arr = ['utm_source', 'utm_medium', 'gh_src', 'utm_term', 'utm_campaign', 'lever-origin']
    for i in split_arr:
        if i in job_url:
            job_url = job_url.split(i)[0]


    conn = sqlite3.connect('vacancies_new.db')
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO vacancies (job_title, job_url, company_name, source_url, added_date, employer_posted_date)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (job_title, job_url, company_name, source_url, added_date, employer_posted_date))
        conn.commit()
        # print(f"Vacancy added: {job_title}")
    except sqlite3.IntegrityError:
        pass
    conn.close()

# Пример добавления вакансии в базу данных
# current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
# add_vacancy("Python Developer", "http://example.com", "Example Company", "http://source.com", "2023-01-01")

# Вывод содержимого таблицы
# display_table()


# Функция для вывода содержимого таблицы
def display_table():
    conn = sqlite3.connect('vacancies.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM vacancies')
    rows = cursor.fetchall()
    conn.close()
    return rows


def count_company_and_vacancies_and_posted_last_week():
    # Определение временного интервала за последние 7 дней
    today = datetime.now()
    seven_days_ago = today - timedelta(days=7)

    # Установка соединения с базой данных
    conn = sqlite3.connect('vacancies.db')
    cursor = conn.cursor()

    # Получение общего количества вакансий
    cursor.execute('SELECT COUNT(*) FROM vacancies')
    total_vacancies = cursor.fetchone()[0]

    # Получение количества уникальных компаний
    cursor.execute('SELECT COUNT(DISTINCT company_name) FROM vacancies WHERE company_name IS NOT NULL')
    unique_company = cursor.fetchone()[0]

    # Получение количества вакансий, опубликованных за последнюю неделю
    cursor.execute("SELECT COUNT(*) FROM vacancies WHERE (employer_posted_date >= ? AND employer_posted_date NOT LIKE '%ago%') OR (CAST(SUBSTR(employer_posted_date, 8, 2) AS INTEGER) < 7 AND employer_posted_date  LIKE '%ago%')", (seven_days_ago.strftime('%Y-%m-%d'), ))

    posted_last_week = cursor.fetchone()[0]

    # Закрытие соединения
    conn.close()

    # Возвращение результатов в виде словаря
    result_dict = {
        'total_vacancies': total_vacancies,
        'unique_company': unique_company,
        'posted_last_week': posted_last_week
    }

    return result_dict


def count_company_and_vacancies():
    conn = sqlite3.connect('vacancies.db')
    cursor = conn.cursor()

    # Получение количества пользователей
    cursor.execute('SELECT COUNT(*) FROM vacancies')
    total_vacancies = cursor.fetchone()[0]

    # Получение количества уникальных ref_id
    cursor.execute('SELECT COUNT(DISTINCT company_name) FROM vacancies WHERE company_name IS NOT NULL')
    unique_company = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM vacancies WHERE DATE(employer_posted_date) >= DATE('now', '-1 day')")
    posted_today = cursor.fetchone()[0]



    # Закрытие соединения
    conn.close()

    # Возвращение результатов в виде словаря
    result_dict = {
        'total_vacancies': total_vacancies,
        'unique_company': unique_company,
        'posted_today':posted_today
    }

    return result_dict

def create_table_db(db_name='vacancies.db'):
    connection = sqlite3.connect(db_name)
    cursor = connection.cursor()

    cursor.execute('''PRAGMA table_info(vacancies)''')
    columns = cursor.fetchall()
    column_names = [column[1] for column in columns]

    if 'categories' not in column_names:
        cursor.execute('''ALTER TABLE vacancies ADD COLUMN categories TEXT''')

    cursor.execute('''SELECT * FROM vacancies''')
    vacancies = cursor.fetchall()

    alias = {
        'Customer Support': ['support', 'customer', 'experience specialist', 'onboarding specialist', 'client', 'help', 'services', 'service', 'success'],
        'Design': ['design', 'graphic', 'ui', 'ux', 'designer', 'artist', 'creative', 'animator', 'illustrator', 'painter', ' art'],
        'Technical': ['engin', 'enginer', 'engineering', 'developer', 'software', 'coder', 'technical', 'researcher', 'data', 'systems', 'tech', ' cto ', 'devops', ' qa', 'security', 'code', 'computer science', 'frontend', 'backend', 'analyst', 'technician', 'cio', 'quantitative', 'fullstack', 'cryptographer', 'architect', 'java', 'ios', 'zkevm', 'cryptographer', ' it ', 'flutter', 'stack', 'react', 'programmer', 'technology'],
        'Finance': ['finance', 'invest', 'risk', 'asset', 'investment', 'treasury', 'financial', 'banking', 'investor', 'trader', 'trade', 'capital', 'audit', 'assets', 'controller', 'regulatory', 'revenue', 'accounting', 'accountant', 'credit', 'investor', 'risk', 'tax', 'liquidity', 'fiat', 'cfo'],
        'Marketing': ['marketing', 'growth', 'social', 'writer', 'community', 'acquisition', 'communications', 'relationships', 'public', 'relations', 'events', 'coordinator', 'content', 'performance', 'campaign', 'video', 'liveops', 'copywriter', 'loyalty', 'advocate', 'relationship', 'aso ', 'seo ', ' pr ', 'ambassador', 'cmo', 'branding', 'media'],
        'Non-Tech': ['non-tech', 'legal', 'talent', ' hr ', 'assistant', 'compliance', 'ta partner', 'policy', 'intern', 'people', 'human', 'paralegal', 'kyc', 'fraud'],
        'Operations': ['manager', 'lead', 'operations', 'vice', 'director', 'president', 'chief', 'associate', 'executive', 'counsel', 'head', ' ops ', 'coordinator', 'strategy', 'officer', 'project', 'producer', 'vp ', 'ceo'],
        'Product': ['product ', 'program manager', ' pm'],
        'Sales': ['business development', 'sale', 'cbdo', 'partnerships', 'relationship', 'account', 'success', 'relationship', 'bdr', 'listing'],
    }
    current_date = datetime.now()
    for vacancy in vacancies:
        job_title = vacancy[1].lower()
        categories = []
        for category, keywords in alias.items():
            for keyword in keywords:
                if keyword in job_title:
                    categories.append(category)
                    break
        if not categories:
            categories.append('Other')
        categories_str = ', '.join(categories)
        cursor.execute('''UPDATE vacancies SET categories = ? WHERE id = ?''', (categories_str, vacancy[0]))
        if('-' not in vacancy[6]):
            result_date = (current_date - timedelta(days=int(re.search(r'\d+', vacancy[6]).group()) )).strftime('%Y-%m-%d')
            cursor.execute('''UPDATE vacancies SET employer_posted_date = ? WHERE id = ?''', (result_date, vacancy[0]))
    connection.commit()
    connection.close()



def create_table_xlsx(title_search, company_search, selected_categories, return_array=False):
    table = display_table()
    types = {
        'Customer Support':[],
        'Design':[],
        'Technical':[],
        'Finance':[],
        'Marketing':[],
        'Non-Tech':[],
        'Operations':[],
        'Product':[],
        'Sales':[],
        'Other':[],
    }

    alias = {
        'Customer Support':['support', 'customer', 'experience specialist', 'onboarding specialist', 'client', 'help', 'services', 'service', 'success'],
        'Design':['design', 'graphic', 'ui', 'ux', 'designer', 'artist', 'creative', 'animator', 'illustrator', 'painter', ' art'],
        'Technical':['engin', 'enginer', 'engineering', 'developer', 'software', 'coder', 'technical', 'researcher', 'data', 'systems', 'tech', ' cto ', 'devops', ' qa', 'security', 'code', 'computer science', 'frontend', 'backend', 'analyst', 'technician', 'cio', 'quantitative', 'fullstack', 'cryptographer', 'architect', 'java', 'ios', 'zkevm', 'cryptographer', ' it ', 'flutter', 'stack', 'react', 'programmer', 'technology'],
        'Finance':['finance', 'invest', 'risk', 'asset', 'investment', 'treasury', 'financial', 'banking', 'investor', 'trader', 'trade', 'capital', 'audit', 'assets', 'controller', 'regulatory', 'revenue', 'accounting', 'accountant', 'credit', 'investor', 'risk', 'tax', 'liquidity', 'fiat', 'cfo'],
        'Marketing':['marketing', 'growth', 'social', 'writer', 'community', 'acquisition', 'communications', 'relationships', 'public', 'relations', 'events', 'coordinator', 'content', 'performance', 'campaign', 'video', 'liveops', 'copywriter', 'loyalty', 'advocate', 'relationship', 'aso ', 'seo ', ' pr ', 'ambassador', 'cmo', 'branding', 'media'],
        'Non-Tech':['non-tech', 'legal', 'talent', ' hr ', 'assistant', 'compliance', 'ta partner', 'policy', 'intern', 'people', 'human', 'paralegal', 'kyc', 'fraud'],
        'Operations':['manager', 'lead', 'operations', 'vice', 'director', 'president', 'chief', 'associate', 'executive', 'counsel', 'head', ' ops ', 'coordinator', 'strategy', 'officer', 'project', 'producer', 'vp ', 'ceo'],
        'Product':['product ', 'program manager', ' pm'],
        'Sales':['business development', 'sale', 'cbdo', 'partnerships', 'relationship', 'account', 'success', 'relationship', 'bdr', 'listing'],
    }

    def write_to_excel(data, file_path='out.xlsx'):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index, value=cell_value)
        workbook.save(file_path)

    if(not return_array):
        data_write_xlsx = []
        for row in table:
            flag=True
            for alias_key in alias:
                for alias_value in alias[alias_key]:
                    if(alias_value in row[1].lower()):
                        types[alias_key].append(row)
                        flag=False
                        data_write_xlsx.append([row[1], alias_key, alias_value])
                        # break
            if(flag):
                types['Other'].append(row)
                data_write_xlsx.append([row[1], 'Other'])


            write_to_excel(data_write_xlsx)
    else:
        data = [['ID','Название вакансии','Ссылка на вакансию','Название компании','Ссылка на источник','Дата добавления','Дата размещения работодателем','Категория','Cлова определения категории']]
        for row in table:
            flag=True
            keys = []
            words = []
            for alias_key in alias:
                for alias_value in alias[alias_key]:
                    if(alias_value in row[1].lower()):
                        types[alias_key].append(row)
                        flag=False
                        if(alias_key not in keys):

                            keys.append(alias_key)
                            words.append(alias_value)
                        # break
            if(flag):
                types['Other'].append(row)
                keys.append('Other')

            if(company_search):
                if(company_search not in row[3]):
                    continue
            if(title_search):
                if(title_search not in row[1]):
                    continue

            if(selected_categories):
                flag_coninue = True
                for cat in selected_categories:
                    if(cat in keys):
                        flag_coninue=False
                if(flag_coninue):
                    continue
            data.append([i for i in row]+[','.join(keys)]+[','.join(words)])

        write_to_excel(data)

def get_daily_statistics(start_date=None, end_date=None):
    try:
        # Проверяем, заданы ли даты начала и конца
        if start_date is None:
            # Если нет, используем предыдущие 30 дней
            start_date = datetime.now() - timedelta(days=30)
        else:
            start_date = datetime.strptime(start_date, "%Y-%m-%d")  # Преобразуем строку в объект datetime
        if end_date is None:
            end_date = datetime.now()
        else:
            end_date = datetime.strptime(end_date, "%Y-%m-%d")  # Преобразуем строку в объект datetime

        # Формируем список для хранения статистики
        daily_statistics = []

        # Подключаемся к базе данных payments.db
        conn_payments = sqlite3.connect('payments.db')
        cursor_payments = conn_payments.cursor()

        # Подключаемся к базе данных users.db
        conn_users = sqlite3.connect('users.db')
        cursor_users = conn_users.cursor()

        # Проходим по каждому дню в указанном диапазоне
        current_date = start_date
        while current_date <= end_date:
            # Форматируем текущую дату в строку
            current_date_str = current_date.strftime("%Y-%m-%d")

            # Получаем выручку за текущий день
            cursor_payments.execute('''
                SELECT SUM(amount)
                FROM payments
                WHERE paid_at IS NOT NULL AND DATE(paid_at) = ?
            ''', (current_date_str,))
            revenue = cursor_payments.fetchone()[0]

            # Получаем новых пользователей за текущий день
            cursor_users.execute('''
                SELECT COUNT(*)
                FROM users
                WHERE registration_date = ?
            ''', (current_date_str,))
            new_users = cursor_users.fetchone()[0]
            # Получаем новых рефералов за текущий день
            cursor_users.execute('''
                SELECT COUNT(DISTINCT ref_id)
                FROM users
                WHERE ref_id IS NOT NULL AND DATE(registration_date) = ?
            ''', (current_date_str,))
            new_referrals = cursor_users.fetchone()[0]

            # Получаем новых клиентов за текущий день
            cursor_users.execute('''
                SELECT COUNT(*)
                FROM users
                WHERE DATE(registration_date) = ? AND ref_id IS NOT NULL
            ''', (current_date_str,))
            new_customers = cursor_users.fetchone()[0]

            # Добавляем данные за текущий день в список статистики
            daily_statistics.append({
                "date": current_date_str,
                "revenue": revenue if revenue else 0,
                "click": new_users if new_users else 0,
                "newReferrals": new_referrals if new_referrals else 0,
                "newCustomers": new_customers if new_customers else 0
            })

            # Переходим к следующему дню
            current_date += timedelta(days=1)

        # Закрываем соединения с базами данных
        conn_payments.close()
        conn_users.close()

        return daily_statistics

    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None

def create_user_counts_table():
    try:
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_counts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                no_paid_count INTEGER DEFAULT 0,
                free_count INTEGER DEFAULT 0,
                promo_count INTEGER DEFAULT 0,
                trial_count INTEGER DEFAULT 0,
                paid_count INTEGER DEFAULT 0
            )
        ''')

        conn.commit()
        conn.close()
    except sqlite3.Error as e:
        print("Ошибка при создании таблицы user_counts:", e)

def get_user_counts(start_date=None, end_date=None):
    try:
        # Проверяем, заданы ли даты начала и конца
        if start_date is None:
            # Если нет, используем предыдущие 30 дней
            start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        if end_date is None:
            end_date = datetime.now().strftime("%Y-%m-%d")

        # Подключаемся к базе данных
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()

        # Получаем данные из таблицы user_counts за указанный период
        cursor.execute('''
            SELECT *
            FROM user_counts
            WHERE date BETWEEN ? AND ?
        ''', (start_date, end_date))
        user_counts_data = cursor.fetchall()

        # Формируем список со словарями данных
        user_counts_list = []
        for row in user_counts_data:
            user_counts_list.append({
                "date": row[1],
                "no_paid": row[2],
                "free": row[3],
                "promo": row[4],
                "trial": row[5],
                "paid": row[6]
            })

        conn.close()

        return user_counts_list

    except sqlite3.Error as e:
        print("Ошибка при получении данных из таблицы user_counts:", e)
        return None

def record_user_counts():
    try:
        # Получаем текущую дату
        current_date = datetime.now().strftime("%Y-%m-%d")

        # Подключаемся к базе данных
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()

        # Проверяем, есть ли уже запись с текущей датой в таблице user_counts
        cursor.execute('SELECT COUNT(*) FROM user_counts WHERE date = ?', (current_date,))
        existing_record = cursor.fetchone()[0]

        if existing_record == 0:
            # Считаем количество пользователей с каждым типом оплаты
            cursor.execute('''
                SELECT 
                    SUM(CASE WHEN subscription_type = 'no paid' THEN 1 ELSE 0 END) AS no_paid_count,
                    SUM(CASE WHEN subscription_type = 'free' THEN 1 ELSE 0 END) AS free_count,
                    SUM(CASE WHEN subscription_type = 'promo' THEN 1 ELSE 0 END) AS promo_count,
                    SUM(CASE WHEN subscription_type = 'trial' THEN 1 ELSE 0 END) AS trial_count,
                    SUM(CASE WHEN subscription_type = 'paid' THEN 1 ELSE 0 END) AS paid_count
                FROM users
            ''')

            # Получаем результат запроса
            counts = cursor.fetchone()

            # Записываем данные о количестве пользователей с каждым типом оплаты и текущей датой
            cursor.execute('''
                INSERT INTO user_counts (date, no_paid_count, free_count, promo_count, trial_count, paid_count)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (current_date, counts[0], counts[1], counts[2], counts[3], counts[4]))

            conn.commit()
        else:
            print("Запись для текущей даты уже существует в таблице user_counts. Новая запись не требуется.")

        conn.close()
        
    except sqlite3.Error as e:
        print("Ошибка при записи данных в таблицу user_counts:", e)

def create_payment_table():
    conn = sqlite3.connect('payments.db')
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            amount REAL,
            external_id TEXT,
            telegram_id INTEGER,
            payment_id TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            paid_at TEXT DEFAULT NULL
        )
    ''')

    conn.commit()
    conn.close()

def add_payment(amount, external_id, telegram_id, payment_id=None):
    conn = sqlite3.connect('payments.db')
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO payments (amount, external_id, telegram_id, payment_id)
        VALUES (?, ?, ?, ?)
    ''', (amount, external_id, telegram_id, payment_id))

    conn.commit()
    conn.close()

def set_payment_date(external_id):
    conn = sqlite3.connect('payments.db')
    cursor = conn.cursor()

    # Получаем текущую дату в формате YYYY-MM-DD
    current_date = datetime.now().strftime('%Y-%m-%d')

    cursor.execute('''
        UPDATE payments
        SET paid_at = ?
        WHERE external_id = ?
    ''', (current_date, external_id))

    conn.commit()

    # Получаем telegram_id по external_id
    cursor.execute('''
        SELECT telegram_id FROM payments
        WHERE external_id = ?
    ''', (external_id,))
    result = cursor.fetchone()
    telegram_id = result[0] if result else None

    conn.close()

    return telegram_id


def get_users_with_expired_subscription():
    # Подключение к базе данных
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # Получение текущей даты
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Получение пользователей с истекшей подпиской
    cursor.execute("SELECT telegram_id, subscription_type FROM users WHERE subscription_end_date <= ? AND subscription_type != 'free' AND subscription_type != 'no paid'", (current_date,))
    expired_users = cursor.fetchall()
    res = [{'telegram_id': user[0], 'subscription_type': user[1]} for user in expired_users]
    # Обновление данных для пользователей с истекшей подпиской
    for user in expired_users:
        telegram_id = user[0]
        update_user_parameter(telegram_id, 'subscription_end_date', '')
        update_user_parameter(telegram_id, 'subscription_type', 'no paid')

    # Закрытие соединения с базой данных
    conn.commit()
    conn.close()

    # Возвращение списка пользователей с истекшей подпиской
    return res

def get_users_subscription_expired_days():
    # Подключение к базе данных
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # Получение текущей даты
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Получение пользователей, у которых подписка истекла 3, 5 или 10 дней назад
    users_expired = []
    cursor.execute("SELECT telegram_id, last_subscription_date FROM users WHERE last_subscription_date IS NOT NULL AND subscription_type == 'no paid' ")
    users = cursor.fetchall()
    for user in users:
        telegram_id, last_subscription_date = user
        if last_subscription_date:
            # Рассчитываем количество дней, прошедших с момента окончания подписки
            days_expired = (datetime.strptime(current_date, '%Y-%m-%d') - datetime.strptime(last_subscription_date, '%Y-%m-%d')).days
            if days_expired in [3, 5, 10]:
                users_expired.append({'telegram_id': telegram_id, 'days_expired': days_expired})

    # Закрытие соединения с базой данных
    conn.close()

    return users_expired

def get_referral_promoters_count(target_telegram_id=None):
    try:
        # Подключаемся к базе данных users.db
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()

        # Формируем запрос для таблицы users
        users_query = '''
            SELECT u1.id, u1.telegram_id, u1.ref_id, u1.registration_date, u1.user_name,
                   COUNT(u2.telegram_id) AS referrals_count,
                   GROUP_CONCAT(u2.telegram_id) AS referral_ids
            FROM users AS u1
            LEFT JOIN users AS u2 ON u1.telegram_id = u2.ref_id
            WHERE u2.ref_id IS NOT NULL
        '''
        if target_telegram_id is not None:
            users_query += ' AND u1.telegram_id = ?'
            cursor.execute(users_query, (target_telegram_id,))
        else:
            cursor.execute(users_query)

        # Получаем данные всех пользователей-промоутеров и количество их приведенных пользователей
        promoter_data = cursor.fetchall()

        # Закрываем соединение с базой данных
        conn.close()

        # Формируем список пользователей-промоутеров с количеством приведенных пользователей
        promoter_list = []
        conn_payments = sqlite3.connect('payments.db')
        cursor_payments = conn_payments.cursor()
        for user in promoter_data:
            user_id, user_telegram_id, ref_id, registration_date, user_name, referrals_count, referral_ids = user
            ref_sum_amout=0
            referral_payments_count=0
            try:
                for lead in referral_ids.split(','):
                    cursor_payments.execute('''
                        SELECT SUM(amount) 
                        FROM payments 
                        WHERE telegram_id=? AND paid_at IS NOT NULL
                    ''', (lead,))
                    total_amount_row = cursor_payments.fetchone()
                    amount = total_amount_row[0] if total_amount_row[0] else 0
                    ref_sum_amout += amount/2
                    if(amount!=0):
                        referral_payments_count += 1
            except:
                pass
            promoter_list.append({
                "id": user_id,
                "telegram_id": user_telegram_id,
                "ref_id": ref_id,
                "registration_date": registration_date,
                "referrals_count": referrals_count,
                "ref_sum_amout": ref_sum_amout,
                "referral_payments_count": referral_payments_count,
                "user_name": user_name
            })

        conn_payments.close()
        return promoter_list

    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None



def get_referral_payments(telegram_id=None):
    try:
        # Подключаемся к базе данных users.db
        conn_users = sqlite3.connect('users.db')
        cursor_users = conn_users.cursor()

        # Формируем запрос для таблицы users
        users_query = '''
            SELECT id, telegram_id, ref_id, registration_date, user_name
            FROM users 
            WHERE ref_id IS NOT ""
        '''
        if telegram_id is not None:
            users_query += ' AND telegram_id = ?'
            cursor_users.execute(users_query, (telegram_id,))
        else:
            cursor_users.execute(users_query)


        # Получаем данные всех пользователей с ref_id != NULL
        users_data = cursor_users.fetchall()

        # Создаем список для хранения данных
        referral_payments = []

        # Подключаемся к базе данных payments.db
        conn_payments = sqlite3.connect('payments.db')
        cursor_payments = conn_payments.cursor()

        # Получаем сумму всех транзакций для каждого пользователя
        for user in users_data:
            user_id, user_telegram_id, ref_id, registration_date, user_name = user
            cursor_payments.execute('''
                SELECT SUM(amount) 
                FROM payments 
                WHERE telegram_id=? AND paid_at IS NOT NULL
            ''', (user_telegram_id,))
            total_amount_row = cursor_payments.fetchone()
            total_amount = total_amount_row[0] if total_amount_row else 0
            
            # Добавляем данные в список
            referral_payments.append({
                "id": user_id,
                "telegram_id": user_telegram_id,
                "ref_id": ref_id,
                "registration_date": registration_date,
                "total_amount": total_amount,
                "user_name": user_name
            })

        # Закрываем соединения с базами данных
        conn_users.close()
        conn_payments.close()

        return referral_payments

    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None

def get_payment_statistics_with_details():
    try:
        # Подключаемся к базе данных payments.db
        conn_payments = sqlite3.connect('payments.db')
        cursor_payments = conn_payments.cursor()

        # Получаем данные о платежах
        cursor_payments.execute('''
            SELECT id, telegram_id, amount, paid_at
            FROM payments
            WHERE paid_at IS NOT NULL
        ''')
        payments_details = cursor_payments.fetchall()

        # Закрываем соединение с базой данных payments.db
        conn_payments.close()

        # Создаем список для хранения деталей платежей
        payments_list = []

        # Переменные для хранения общей суммы платежей и уникальных пользователей
        total_payments_sum = 0
        unique_users = set()
        paid_20_count = 0
        paid_200_count = 0

        # Подключаемся к базе данных users.db
        conn_users = sqlite3.connect('users.db')
        cursor_users = conn_users.cursor()

        # Для каждого платежа получаем данные о пользователе из users.db
        for payment in payments_details:
            payment_id, telegram_id, amount, paid_at = payment

            # Получаем информацию о пользователе из таблицы users
            cursor_users.execute('SELECT user_name, ref_id FROM users WHERE telegram_id=?', (telegram_id,))
            user_info = cursor_users.fetchone()
            user_name = user_info[0] if user_info else "Unknown User"
            ref_id = user_info[1] if user_info else None

            # Обновляем сумму всех платежей и добавляем уникальных пользователей
            total_payments_sum += amount
            unique_users.add(telegram_id)

            # Проверяем сумму платежа и увеличиваем счетчики соответственно
            if amount >= 200:
                paid_200_count += 1
            elif amount >= 20:
                paid_20_count += 1

            # Добавляем информацию о платеже в список
            payments_list.append({
                "id": payment_id,
                "telegram_id": telegram_id,
                "user_name": user_name,
                "amount": amount,
                "ref_id": ref_id,
                "paid_at": paid_at
            })

        # Закрываем соединение с базой данных users.db
        conn_users.close()

        # Получаем общее количество уникальных пользователей
        unique_users_count = len(unique_users)

        return {
            "payments_list": payments_list,
            "total_payments_sum": total_payments_sum,
            "unique_users_count": unique_users_count,
            "paid_20_count": paid_20_count,
            "paid_200_count": paid_200_count
        }

    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None


def create_users_table():
    # Подключение к базе данных (если её нет, она будет создана)
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # Создание таблицы users (если она не существует)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE NOT NULL,
            registration_date TEXT NOT NULL,
            subscription_end_date TEXT,
            subscription_type TEXT DEFAULT 'free',
            category TEXT DEFAULT '',
            keywords TEXT DEFAULT '',
            companies TEXT DEFAULT '',
            balance INTEGER DEFAULT 0,
            ref_id INTEGER,
            notification INTEGER DEFAULT 1,
            last_subscription_date TEXT DEFAULT '',
            wallet TEXT DEFAULT '',
            user_name TEXT DEFAULT ''
        )
    ''')

    # Сохранение изменений и закрытие соединения
    conn.commit()
    conn.close()

def create_transactions_table():
    try:
        conn = sqlite3.connect('users.db')
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                transaction_id TEXT,
                datetime TEXT,
                amount REAL,
                wallet TEXT,
                telegram_id INTEGER
            )
        ''')
        conn.commit()
        conn.close()
    except sqlite3.Error as e:
        pass

def get_transaction_info(transaction_id):
    try:
        conn = sqlite3.connect('users.db')
        c = conn.cursor()
        c.execute('''
            SELECT * FROM transactions WHERE id = ?
        ''', (transaction_id,))
        result = c.fetchone()
        conn.close()
        if result:
            transaction_info = {
                "id": result[0],
                "transaction_id": result[1],
                "datetime": result[2],
                "amount": result[3],
                "wallet": result[4],
                "telegram_id": result[5]
            }
            return transaction_info
        else:
            return None
    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None


def get_transactions_without_id():
    try:
        conn = sqlite3.connect('users.db')
        c = conn.cursor()

        # Получаем список всех транзакций без идентификатора транзакции
        c.execute('SELECT * FROM transactions WHERE transaction_id IS NULL')
        transactions = c.fetchall()

        # Получаем количество таких транзакций
        transactions_count = len(transactions)

        # Получаем общую сумму всех таких транзакций
        total_amount = sum(transaction[3] for transaction in transactions)


        c.execute('SELECT COUNT(DISTINCT telegram_id) FROM transactions WHERE transaction_id IS NULL')
        unique_users_count = c.fetchone()[0]
        c.close()

        return {
            "transactions": transactions,
            "transactions_count": transactions_count,
            "total_amount": total_amount,
            "unique_users_count": unique_users_count
        }
    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None

def get_transactions_with_id():
    try:
        conn = sqlite3.connect('users.db')
        c = conn.cursor()

        # Получаем список всех транзакций без идентификатора транзакции
        c.execute('SELECT * FROM transactions WHERE transaction_id IS NOT NULL')
        transactions = c.fetchall()

        # Получаем количество таких транзакций
        transactions_count = len(transactions)

        # Получаем общую сумму всех таких транзакций
        total_amount = sum(transaction[3] for transaction in transactions)


        # Получаем количество уникальных пользователей по telegram_id
        c.execute('SELECT COUNT(DISTINCT telegram_id) FROM transactions WHERE transaction_id IS NOT NULL')
        unique_users_count = c.fetchone()[0]
        c.close()

        return {
            "transactions": transactions,
            "transactions_count": transactions_count,
            "total_amount": total_amount,
            "unique_users_count": unique_users_count
        }
    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None


def update_transaction_id_by_id(id, new_transaction_id):
    try:
        conn = sqlite3.connect('users.db')
        c = conn.cursor()

        # Обновляем transaction_id для указанного id
        c.execute('UPDATE transactions SET transaction_id=? WHERE id=?', (new_transaction_id, id))
        conn.commit()

        conn.close()

        return True
    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return False

def count_paid_referrals(telegram_id):
    try:
        conn = sqlite3.connect('users.db')
        c = conn.cursor()

        # Получаем всех пользователей, у которых ref_id равен переданному telegram_id
        c.execute('''
            SELECT telegram_id FROM users WHERE ref_id=?
        ''', (telegram_id,))
        
        # Получение списка всех рефералов
        referrals = c.fetchall()
        
        conn.close()
        conn = sqlite3.connect('payments.db')
        c = conn.cursor()
        # Счетчик оплаченных рефералов
        paid_referrals_count = 0

        # Проверяем для каждого реферала, есть ли запись в таблице payments, где paid_at не равен NULL
        for referral in referrals:
            c.execute('''
                SELECT COUNT(*) FROM payments WHERE telegram_id=? AND paid_at IS NOT NULL
            ''', (referral[0],))
            count = c.fetchone()[0]
            if count > 0:
                paid_referrals_count += 1

        conn.close()
        
        return paid_referrals_count

    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None

def get_user_payment_info(telegram_id):
    try:
        conn = sqlite3.connect('users.db')
        c = conn.cursor()

        c.execute('SELECT SUM(amount) FROM transactions WHERE telegram_id=? AND transaction_id IS NOT NULL', (telegram_id,))
        total_amount_with_id_row = c.fetchone()
        total_amount_with_id = total_amount_with_id_row[0] if total_amount_with_id_row[0] else 0

        # Получаем сумму всех платежей пользователя из таблицы transactions, где нет payment_id
        c.execute('SELECT SUM(amount) FROM transactions WHERE telegram_id=? AND transaction_id IS NULL', (telegram_id,))
        total_amount_without_id_row = c.fetchone()
        total_amount_without_id = total_amount_without_id_row[0] if total_amount_without_id_row[0] else 0

        conn.close()

        return {
            "total_amount_with_id": total_amount_with_id,
            "total_amount_without_id": total_amount_without_id
        }
    except sqlite3.Error as e:
        print("Ошибка при выполнении запроса к базе данных:", e)
        return None



# print(get_user_payment_info(5318713823))

def create_transactions_xlsx(telegram_id):
    db_file = 'users.db'
    xlsx_filename = f'{telegram_id}.xlsx'

    try:
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        
        # Получаем транзакции для данного telegram_id, отсортированные по убыванию даты
        c.execute('''
            SELECT * FROM transactions WHERE telegram_id=? ORDER BY datetime DESC
        ''', (telegram_id,))
        transactions = c.fetchall()
        
        # Создаем новый xlsx файл
        workbook = Workbook()
        worksheet = workbook.active

        # Записываем заголовки столбцов
        headers = ['transaction_id', 'datetime', 'amount', 'wallet', 'telegram_id']
        worksheet.append(headers)

        # Записываем транзакции
        for transaction in transactions:
            worksheet.append(transaction[1:])

        # Сохраняем файл
        workbook.save(filename=xlsx_filename)
        conn.close()

        return xlsx_filename

    except sqlite3.Error as e:
        return None



def add_transaction(amount, wallet, telegram_id, transaction_id=None):
    conn = sqlite3.connect('users.db')
    c = conn.cursor()
    if transaction_id:
        c.execute('''
            INSERT INTO transactions (transaction_id, datetime, amount, wallet, telegram_id)
            VALUES (?, ?, ?, ?, ?)
        ''', (transaction_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), amount, wallet, telegram_id))
    else:
        c.execute('''
            INSERT INTO transactions (datetime, amount, wallet, telegram_id)
            VALUES (?, ?, ?, ?)
        ''', (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), amount, wallet, telegram_id))
    conn.commit()
    conn.close()

def count_referrals(telegram_id):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # Получение количества пользователей с указанным ref_id
    cursor.execute('SELECT COUNT(*) FROM users WHERE ref_id = ?', (telegram_id,))
    count = cursor.fetchone()[0]

    # Закрытие соединения
    conn.close()

    return count

def create_user(telegram_id, ref_id=0):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # Проверка, есть ли уже пользователь с таким telegram_id
    cursor.execute('SELECT * FROM users WHERE telegram_id = ?', (telegram_id,))
    existing_user = cursor.fetchone()

    # Если пользователь уже существует, ничего не делаем
    if existing_user:
        conn.close()
        return True

    # Получение текущей даты в формате YYYY-MM-DD
    registration_date = datetime.now().strftime('%Y-%m-%d')

    # Вставка нового пользователя
    cursor.execute('''
        INSERT INTO users (telegram_id, registration_date, ref_id)
        VALUES (?, ?, ?)
    ''', (telegram_id, registration_date, ref_id))

    # Сохранение изменений и закрытие соединения
    conn.commit()
    conn.close()
    return False

def get_user_categories(telegram_id):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # Получение данных пользователя по telegram_id
    cursor.execute('SELECT category FROM users WHERE telegram_id = ?', (telegram_id,))
    user_categories = cursor.fetchone()

    # Закрытие соединения
    conn.close()

    if user_categories:
        # Разделение категорий по запятой и возвращение в виде списка
        categories_list = user_categories[0].split(', ')
        return categories_list
    else:
        return []

def get_user_data(telegram_id):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # Получение данных пользователя по telegram_id
    cursor.execute('SELECT * FROM users WHERE telegram_id = ?', (telegram_id,))
    user_data = cursor.fetchone()

    # Закрытие соединения
    conn.close()

    if user_data:
        # Преобразование кортежа в словарь для удобства
        user_dict = {
            'id': user_data[0],
            'telegram_id': user_data[1],
            'registration_date': user_data[2],
            'subscription_end_date': user_data[3],
            'subscription_type': user_data[4],
            'category': user_data[5],
            'keywords': user_data[6],
            'companies': user_data[7],
            'balance': user_data[8],
            'ref_id': user_data[9],
            'notification': user_data[10],
            'last_subscription_date': user_data[11],
            'wallet': user_data[12],
            'user_name': user_data[13]
        }
        return user_dict
    else:
        return None

def get_user_table():
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users')
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_user_with_parameter(parameter, value):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f'SELECT telegram_id FROM users where {parameter} = ? AND notification = ?', (value, 1))
    rows = cursor.fetchall()
    conn.close()
    return rows

def update_user_parameter(telegram_id, parameter, value):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # Обновление параметра пользователя
    cursor.execute(f'UPDATE users SET {parameter} = ? WHERE telegram_id = ?', (value, telegram_id))

    # Сохранение изменений и закрытие соединения
    conn.commit()
    conn.close()

def set_month_promo(telegram_id):
    user = get_user_data(telegram_id)
    if(user['subscription_end_date']==None or user['subscription_end_date']==''):
        current_date = datetime.now()
        new_date = current_date + timedelta(days=30)
        new_date = new_date.strftime("%Y-%m-%d")
    else:
        subscription_end_date = user['subscription_end_date']
        date_obj = datetime.strptime(subscription_end_date, "%Y-%m-%d")
        new_date_obj = date_obj + timedelta(days=30)
        new_date = new_date_obj.strftime("%Y-%m-%d")

    update_user_parameter(telegram_id, 'subscription_end_date', new_date)
    update_user_parameter(telegram_id, 'subscription_type', 'promo')
    update_user_parameter(telegram_id, 'last_subscription_date', new_date)


def create_msg_table():
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            paid_year BOOLEAN,
            paid_monthly BOOLEAN,
            trial BOOLEAN,
            promo BOOLEAN,
            free BOOLEAN,
            msg TEXT,
            image_path TEXT
        )
    ''')

    conn.commit()
    conn.close()

def add_message(paid_year, paid_monthly, trial, promo, free, msg, image_path):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO messages (paid_year, paid_monthly, trial, promo, free, msg, image_path)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (paid_year, paid_monthly, trial, promo, free, msg, image_path))

    conn.commit()
    conn.close()
    print('Записанно')

def delete_earliest_message(message_id):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    cursor.execute('''
        DELETE FROM messages WHERE id = ?
    ''', (message_id,))

    conn.commit()
    conn.close()

def get_earliest_message():
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    cursor.execute('''
        SELECT * FROM messages ORDER BY id ASC LIMIT 1
    ''')

    result = cursor.fetchone()

    conn.close()

    if result:
        message_dict = {
            'paid_year': bool(result[1]),
            'paid_monthly': bool(result[2]),
            'trial': bool(result[3]),
            'promo': bool(result[4]),
            'free': bool(result[5]),
            'msg': result[6],
            'image_path': result[7]
        }

        # Удаляем самую раннюю запись
        delete_earliest_message(result[0])

        return message_dict
    else:
        return None

def count_new_vacancies_by_category():
    conn_vacancies = sqlite3.connect('vacancies.db')
    cursor_vacancies = conn_vacancies.cursor()

    conn_vacancies_new = sqlite3.connect('vacancies_new.db')
    cursor_vacancies_new = conn_vacancies_new.cursor()

    cursor_vacancies.execute("SELECT categories FROM vacancies;")
    all_categories = cursor_vacancies.fetchall()

    individual_categories = set()
    for categories in all_categories:
        individual_categories.update(categories[0].split(', '))

    new_vacancies_by_category = {category: {"count": 0, "vacancies": []} for category in individual_categories}

    for category in individual_categories:
        query = "SELECT job_title, job_url, company_name FROM vacancies WHERE categories LIKE ?;"
        cursor_vacancies_new.execute(query, ('%'+category+'%',))
        new_job_data = cursor_vacancies_new.fetchall()

        for job_title, job_url, company in new_job_data:
            cursor_vacancies.execute("SELECT job_url FROM vacancies WHERE job_url = ?;", (job_url,))
            if cursor_vacancies.fetchone() is None:
                new_vacancies_by_category[category]["count"] += 1
                new_vacancies_by_category[category]["vacancies"].append({"title": job_title, "url": job_url, "company": company})

    conn_vacancies.close()
    conn_vacancies_new.close()

    return new_vacancies_by_category

def get_users_new_vacancies():
    # Подключаемся к базе данных пользователей
    conn_users = sqlite3.connect('users.db')
    cursor_users = conn_users.cursor()

    # Получаем всех пользователей с включенными уведомлениями и их категории
    cursor_users.execute("""
        SELECT telegram_id, category 
        FROM users 
        WHERE notification = 1;
    """)
    users_with_notifications = cursor_users.fetchall()

    conn_users.close()

    # Получаем все новые вакансии по категориям заранее
    new_vacancies = count_new_vacancies_by_category()

    users_new_vacancies = []

    for telegram_id, user_categories in users_with_notifications:
        user_vacancies = []
        
        # Создаем множество для хранения уже добавленных вакансий
        added_urls = set()
        
        # Если у пользователя нет категорий, добавляем все новые вакансии
        if not user_categories:
            for category_vacancies in new_vacancies.values():
                for vacancy in category_vacancies["vacancies"]:
                    if vacancy["url"] not in added_urls:
                        user_vacancies.append(vacancy)
                        added_urls.add(vacancy["url"])
        else:
            # Если у пользователя есть категории, добавляем только соответствующие вакансии
            categories = user_categories.split(', ')
            for category in categories:
                category_vacancies = new_vacancies.get(category, {"vacancies": []})
                for vacancy in category_vacancies["vacancies"]:
                    if vacancy["url"] not in added_urls:
                        vacancy["category"] = category  # Добавляем категорию к вакансии
                        user_vacancies.append(vacancy)
                        added_urls.add(vacancy["url"])

        users_new_vacancies.append({"id": telegram_id, "vacancies": user_vacancies})

    return users_new_vacancies


# Пример использования функции

create_msg_table()
# Вызов функции для создания таблицы
create_users_table()
# Создание таблицы, если её ещё нет
create_table()
create_table_db()

# Создаем таблицу платежей
create_payment_table()

create_transactions_table()
create_user_counts_table()